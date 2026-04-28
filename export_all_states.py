"""
Export Holt-Winters forecasts for ALL US states to JSON.
Run this once before starting the Phoenix app:
    python export_all_states.py
Output: medicaid_forecast/priv/data/all_states_forecast.json
"""

import glob
import json
import os
import re
import warnings
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import mean_absolute_error, mean_squared_error
from statsmodels.tsa.holtwinters import ExponentialSmoothing

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

NUMERIC_COLS = [
    "Total Investigations", "Fraud Investigations", "Abuse/Neglect Investigations",
    "Total Indictments", "Fraud Indictments", "Abuse/Neglect Indictments",
    "Total Convictions", "Fraud Convictions", "Abuse/Neglect Convictions",
    "Civil Settlements and Judgments", "Total Recoveries",
    "Total Criminal Recoveries", "Civil Recoveries Global", "Civil Recoveries Other",
    "MFCU Grant Expenditures", "Total Medicaid Expenditures", "Staff On Board",
]

SKIP_NAMES = {"total", "grand total", ""}
SPLIT_YEAR = 2022
FORECAST_YEARS = list(range(2026, 2036))
N_FORECAST = len(FORECAST_YEARS)
MIN_DATA_POINTS = 5

OUT_DIR = os.path.join("medicaid_forecast", "priv", "data")
OUT_FILE = os.path.join(OUT_DIR, "all_states_forecast.json")

# ---------------------------------------------------------------------------
# Data loader (same logic as all_models.py)
# ---------------------------------------------------------------------------

def load_year(filepath):
    year = int(re.search(r"FY_(\d{4})", filepath).group(1))
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    ncols = ws.max_column

    def to_num(v):
        return v if isinstance(v, (int, float)) else None

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        state = row[0]
        if state is None or not isinstance(state, str) or len(state) > 40:
            continue
        state_clean = state.strip()
        if state_clean.upper().replace(":", "").strip() in ("TOTAL", "GRAND TOTAL"):
            continue  # skip aggregate rows

        if ncols == 17:
            nums = [to_num(row[i]) for i in range(1, 14)] + [None] + [to_num(row[i]) for i in range(14, 17)]
        else:
            nums = [to_num(row[i]) for i in range(1, 18)]
        rows.append([year, state_clean] + nums)

    return pd.DataFrame(rows, columns=["FY", "State"] + NUMERIC_COLS)


# ---------------------------------------------------------------------------
# Load all files
# ---------------------------------------------------------------------------

files = sorted(glob.glob("FY_*_MFCU_Statistical_Chart*.xlsx"))
print(f"Loading {len(files)} Excel files...")
all_data = pd.concat([load_year(f) for f in files], ignore_index=True)

# Normalize state names: strip whitespace, title-case
all_data["State"] = all_data["State"].str.strip().str.title()

# Get sorted unique state names (exclude aggregate rows)
skip_set = {"Total", "Grand Total", ""}
state_names = sorted(
    s for s in all_data["State"].unique()
    if s not in skip_set and len(s) <= 35
)
print(f"Found {len(state_names)} unique state/territory names")

historical_years = sorted(all_data["FY"].unique().tolist())

# ---------------------------------------------------------------------------
# Run Holt-Winters for each state
# ---------------------------------------------------------------------------

state_data = {}
skipped = []

for state in state_names:
    rows = all_data[all_data["State"] == state].sort_values("FY")
    rows = rows.dropna(subset=["Total Medicaid Expenditures"]).reset_index(drop=True)

    if len(rows) < MIN_DATA_POINTS:
        skipped.append(f"{state} ({len(rows)} pts)")
        continue

    years = rows["FY"].values
    y = rows["Total Medicaid Expenditures"].values / 1e9  # $Billions

    # Align to full historical_years range (fill missing years with NaN for table only)
    hist_map = dict(zip(years.tolist(), y.tolist()))
    historical_vals = [hist_map.get(yr) for yr in historical_years]

    # Train/test split
    train_mask = years <= SPLIT_YEAR
    y_train = y[train_mask]
    y_test = y[~train_mask]
    n_test = len(y_test)

    if len(y_train) < 3:
        skipped.append(f"{state} (insufficient train data)")
        continue

    try:
        # Train-only model for test metrics
        hw_train = ExponentialSmoothing(y_train, trend="add", seasonal=None).fit(optimized=True)
        fc_test = hw_train.forecast(n_test)

        # Full model for forecasting
        hw_full = ExponentialSmoothing(y, trend="add", seasonal=None).fit(optimized=True)
        fc_future = hw_full.forecast(N_FORECAST)

        # 95% CI (bootstrap from residuals)
        resid_std = float(np.std(hw_full.resid))
        h_arr = np.arange(1, N_FORECAST + 1)
        ci_lower = fc_future - 1.96 * resid_std * np.sqrt(h_arr)
        ci_upper = fc_future + 1.96 * resid_std * np.sqrt(h_arr)

        # Test metrics (only if we have test data)
        metrics = {}
        if n_test > 0 and len(fc_test) == n_test:
            mae = float(mean_absolute_error(y_test, fc_test))
            rmse = float(np.sqrt(mean_squared_error(y_test, fc_test)))
            with np.errstate(divide="ignore", invalid="ignore"):
                mape_vals = np.abs((y_test - fc_test) / y_test)
                mape = float(np.nanmean(mape_vals) * 100)
            metrics = {"mae": round(mae, 4), "rmse": round(rmse, 4), "mape": round(mape, 2)}

        def clean(arr):
            return [round(float(v), 4) if v is not None and not np.isnan(float(v)) else None
                    for v in arr]

        state_data[state] = {
            "historical": clean(historical_vals),
            "forecast":   clean(fc_future),
            "ci_lower":   clean(ci_lower),
            "ci_upper":   clean(ci_upper),
            "metrics":    metrics,
        }

    except Exception as e:
        skipped.append(f"{state} (error: {e})")

print(f"Computed forecasts for {len(state_data)} states.")
if skipped:
    print(f"Skipped {len(skipped)}: {', '.join(skipped)}")

# ---------------------------------------------------------------------------
# Write JSON
# ---------------------------------------------------------------------------

os.makedirs(OUT_DIR, exist_ok=True)

output = {
    "generated_at": str(date.today()),
    "historical_years": historical_years,
    "forecast_years": FORECAST_YEARS,
    "states": sorted(state_data.keys()),
    "data": state_data,
}

with open(OUT_FILE, "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2)

print(f"\nSaved → {OUT_FILE}")
print(f"States included: {len(state_data)}")
print(f"Historical years: {historical_years[0]}–{historical_years[-1]}")
print(f"Forecast years:   {FORECAST_YEARS[0]}–{FORECAST_YEARS[-1]}")
