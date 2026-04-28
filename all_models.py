"""
Medicaid Expenditure Projection — All Four Models
Models: Linear/Polynomial Regression, ARIMA, Holt-Winters, Prophet
Forecasts NY Total Medicaid Expenditures FY 2026-2035
Outputs a comparison table with MAE, RMSE, MAPE and a combined chart.
"""

import glob
import re
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.pipeline import make_pipeline
from sklearn.metrics import mean_absolute_error, mean_squared_error
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from prophet import Prophet

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# 1.  Load & filter NY data  (same loader as regression_model.py)
# ---------------------------------------------------------------------------

NUMERIC_COLS = [
    "Total Investigations", "Fraud Investigations", "Abuse/Neglect Investigations",
    "Total Indictments", "Fraud Indictments", "Abuse/Neglect Indictments",
    "Total Convictions", "Fraud Convictions", "Abuse/Neglect Convictions",
    "Civil Settlements and Judgments", "Total Recoveries",
    "Total Criminal Recoveries", "Civil Recoveries Global", "Civil Recoveries Other",
    "MFCU Grant Expenditures", "Total Medicaid Expenditures", "Staff On Board",
]
NY_ALIASES = {"new york", "new york state", "ny"}


def load_year(filepath):
    year = int(re.search(r"FY_(\d{4})", filepath).group(1))
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    ncols = ws.max_column

    def to_num(v):
        return v if isinstance(v, (int, float)) else None

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        state = row[0]
        if state is None or not isinstance(state, str) or len(state) > 30:
            continue
        if state.strip().upper().replace(":", "").strip() in ("TOTAL", "GRAND TOTAL"):
            state = "Total"
        if ncols == 17:
            nums = [to_num(row[i]) for i in range(1, 14)] + [None] + [to_num(row[i]) for i in range(14, 17)]
        else:
            nums = [to_num(row[i]) for i in range(1, 18)]
        rows.append([year, state] + nums)

    return pd.DataFrame(rows, columns=["FY", "State"] + NUMERIC_COLS)


files = sorted(glob.glob("FY_*_MFCU_Statistical_Chart*.xlsx"))
all_data = pd.concat([load_year(f) for f in files], ignore_index=True)
ny = all_data[all_data["State"].str.strip().str.lower().isin(NY_ALIASES)].copy()
ny = ny.sort_values("FY").reset_index(drop=True)
ts = ny.dropna(subset=["Total Medicaid Expenditures"]).reset_index(drop=True)

years = ts["FY"].values
y_raw = ts["Total Medicaid Expenditures"].values          # original dollars
y = y_raw / 1e9                                           # billions

print(f"Loaded {len(files)} files | {len(ts)} NY rows: FY {years[0]}–{years[-1]}")

# ---------------------------------------------------------------------------
# 2.  Train / test split  (train: 2013-2022, test: 2023-2025)
# ---------------------------------------------------------------------------

SPLIT_YEAR = 2022
FORECAST_YEARS = list(range(2026, 2036))   # 10-year forecast (2026-2035)
N_FORECAST = len(FORECAST_YEARS)

train_mask = years <= SPLIT_YEAR
y_train, y_test = y[train_mask], y[~train_mask]
yrs_train, yrs_test = years[train_mask], years[~train_mask]

n_test = len(y_test)

# ---------------------------------------------------------------------------
# 3.  Model A — Polynomial Regression (best of deg 1/2/3 by train R²)
# ---------------------------------------------------------------------------

best_poly, best_r2_train = None, -np.inf
for deg in [1, 2, 3]:
    pipe = make_pipeline(PolynomialFeatures(deg), LinearRegression())
    pipe.fit(yrs_train.reshape(-1, 1), y_train)
    fitted = pipe.predict(yrs_train.reshape(-1, 1))
    ss_res = np.sum((y_train - fitted) ** 2)
    ss_tot = np.sum((y_train - np.mean(y_train)) ** 2)
    r2 = 1 - ss_res / ss_tot
    if r2 > best_r2_train:
        best_poly, best_r2_train, best_deg = pipe, r2, deg

pred_test_reg = best_poly.predict(yrs_test.reshape(-1, 1))
pred_future_reg = best_poly.predict(np.array(FORECAST_YEARS).reshape(-1, 1))

mae_reg  = mean_absolute_error(y_test, pred_test_reg)
rmse_reg = np.sqrt(mean_squared_error(y_test, pred_test_reg))
mape_reg = np.mean(np.abs((y_test - pred_test_reg) / y_test)) * 100

# in-sample fitted (all historical years, for plot)
y_fit_reg = best_poly.predict(years.reshape(-1, 1))

print(f"\n[A] Polynomial Regression (deg {best_deg}) — test MAE={mae_reg:.3f}B  RMSE={rmse_reg:.3f}B  MAPE={mape_reg:.2f}%")

# ---------------------------------------------------------------------------
# 4.  Model B — ARIMA
# ---------------------------------------------------------------------------

# Auto-search simple ARIMA orders on training data
from itertools import product as iproduct

best_arima, best_aic = None, np.inf
for p, d, q in iproduct(range(3), range(2), range(3)):
    try:
        m = ARIMA(y_train, order=(p, d, q)).fit()
        if m.aic < best_aic:
            best_aic, best_arima = m.aic, m
            best_pdq = (p, d, q)
    except Exception:
        pass

# Re-fit on full history for forecasting
arima_full = ARIMA(y, order=best_pdq).fit()

# Test predictions: use model fitted on train, forecast n_test steps
arima_train_fit = best_arima
fc_test_arima = arima_train_fit.forecast(steps=n_test)
fc_future_arima = arima_full.forecast(steps=N_FORECAST)

mae_arima  = mean_absolute_error(y_test, fc_test_arima)
rmse_arima = np.sqrt(mean_squared_error(y_test, fc_test_arima))
mape_arima = np.mean(np.abs((y_test - fc_test_arima) / y_test)) * 100

# 95% CI for forecast
_arima_fc_obj = arima_full.get_forecast(steps=N_FORECAST)
fc_arima_ci = _arima_fc_obj.conf_int(alpha=0.05)
# conf_int() returns a DataFrame; extract as arrays
arima_ci_lower_arr = fc_arima_ci.iloc[:, 0].values if hasattr(fc_arima_ci, "iloc") else fc_arima_ci[:, 0]
arima_ci_upper_arr = fc_arima_ci.iloc[:, 1].values if hasattr(fc_arima_ci, "iloc") else fc_arima_ci[:, 1]

print(f"[B] ARIMA{best_pdq}          — test MAE={mae_arima:.3f}B  RMSE={rmse_arima:.3f}B  MAPE={mape_arima:.2f}%")

# ---------------------------------------------------------------------------
# 5.  Model C — Holt-Winters Exponential Smoothing
# ---------------------------------------------------------------------------

# Annual data → no true seasonality; use additive trend
hw_train = ExponentialSmoothing(y_train, trend="add", seasonal=None).fit(optimized=True)
fc_test_hw = hw_train.forecast(n_test)

hw_full = ExponentialSmoothing(y, trend="add", seasonal=None).fit(optimized=True)
fc_future_hw = hw_full.forecast(N_FORECAST)

mae_hw  = mean_absolute_error(y_test, fc_test_hw)
rmse_hw = np.sqrt(mean_squared_error(y_test, fc_test_hw))
mape_hw = np.mean(np.abs((y_test - fc_test_hw) / y_test)) * 100

# Bootstrap-style 95% CI for Holt-Winters (simulated from residuals)
hw_resid_std = np.std(hw_full.resid)
hw_ci_lower = fc_future_hw - 1.96 * hw_resid_std * np.sqrt(np.arange(1, N_FORECAST + 1))
hw_ci_upper = fc_future_hw + 1.96 * hw_resid_std * np.sqrt(np.arange(1, N_FORECAST + 1))

print(f"[C] Holt-Winters           — test MAE={mae_hw:.3f}B  RMSE={rmse_hw:.3f}B  MAPE={mape_hw:.2f}%")

# ---------------------------------------------------------------------------
# 6.  Model D — Prophet
# ---------------------------------------------------------------------------

# Prophet expects a DataFrame with columns ds (date) and y
def make_prophet_df(yrs, vals):
    return pd.DataFrame({"ds": pd.to_datetime([f"{y}-07-01" for y in yrs]), "y": vals})

prophet_train_df = make_prophet_df(yrs_train, y_train)
m_prophet_train = Prophet(yearly_seasonality=False, weekly_seasonality=False,
                          daily_seasonality=False, interval_width=0.95)
m_prophet_train.fit(prophet_train_df)

future_test_df = make_prophet_df(yrs_test, y_test)
fc_test_prophet = m_prophet_train.predict(future_test_df)["yhat"].values

# Full model for 10-year forecast
prophet_full_df = make_prophet_df(years, y)
m_prophet_full = Prophet(yearly_seasonality=False, weekly_seasonality=False,
                         daily_seasonality=False, interval_width=0.95)
m_prophet_full.fit(prophet_full_df)

future_fc_df = make_prophet_df(FORECAST_YEARS, [0] * N_FORECAST)
prophet_fc = m_prophet_full.predict(future_fc_df)
fc_future_prophet = prophet_fc["yhat"].values
prophet_ci_lower  = prophet_fc["yhat_lower"].values
prophet_ci_upper  = prophet_fc["yhat_upper"].values

mae_prophet  = mean_absolute_error(y_test, fc_test_prophet)
rmse_prophet = np.sqrt(mean_squared_error(y_test, fc_test_prophet))
mape_prophet = np.mean(np.abs((y_test - fc_test_prophet) / y_test)) * 100

print(f"[D] Prophet                — test MAE={mae_prophet:.3f}B  RMSE={rmse_prophet:.3f}B  MAPE={mape_prophet:.2f}%")

# ---------------------------------------------------------------------------
# 7.  Summary comparison table
# ---------------------------------------------------------------------------

models_meta = [
    ("Poly Regression", f"deg {best_deg}", mae_reg,     rmse_reg,     mape_reg,     pred_future_reg),
    ("ARIMA",           str(best_pdq),     mae_arima,   rmse_arima,   mape_arima,   fc_future_arima),
    ("Holt-Winters",    "add trend",       mae_hw,      rmse_hw,      mape_hw,      fc_future_hw),
    ("Prophet",         "linear trend",    mae_prophet, rmse_prophet, mape_prophet, fc_future_prophet),
]

print("\n" + "=" * 75)
print("MODEL PERFORMANCE COMPARISON  (test set FY 2023–2025, values in $Billions)")
print("=" * 75)
print(f"{'Model':<18} {'Config':<15} {'MAE':>8} {'RMSE':>8} {'MAPE':>8}")
print("-" * 60)
for name, cfg, mae, rmse, mape, _ in models_meta:
    print(f"{name:<18} {cfg:<15} {mae:>7.3f}B {rmse:>7.3f}B {mape:>7.2f}%")

print("\n" + "=" * 95)
print("10-YEAR FORECAST  —  NY Total Medicaid Expenditures ($Billions)")
print("=" * 95)
header = f"{'Year':<6}" + "".join(f"{name:>18}" for name, *_ in models_meta)
print(header)
print("-" * 78)
for i, yr in enumerate(FORECAST_YEARS):
    row = f"{yr:<6}" + "".join(f"${m[-1][i]:>15.2f}B" for m in models_meta)
    print(row)

# ---------------------------------------------------------------------------
# 8.  Visualization — redesigned 2×2 grid
# ---------------------------------------------------------------------------

# ── Style & palette ──────────────────────────────────────────────────────────
plt.rcParams.update({
    "font.family":      "DejaVu Sans",
    "axes.spines.top":  False,
    "axes.spines.right":False,
    "axes.grid":        True,
    "grid.color":       "#E0E0E0",
    "grid.linewidth":   0.7,
    "axes.titlesize":   12,
    "axes.titleweight": "bold",
    "axes.labelsize":   10,
    "xtick.labelsize":  9,
    "ytick.labelsize":  9,
    "legend.fontsize":  8.5,
    "legend.framealpha":0.9,
    "legend.edgecolor": "#CCCCCC",
})

C = {
    "Poly Regression": "#2E86AB",   # ocean blue
    "ARIMA":           "#E07B54",   # terracotta
    "Holt-Winters":    "#28A745",   # forest green
    "Prophet":         "#8E44AD",   # purple
    "actual":          "#1C1C1C",
    "divider":         "#9E9E9E",
    "history":         "#CFE2FF",
}

DOLLAR_FMT = mticker.FuncFormatter(lambda x, _: f"${x:,.0f}B")

fig = plt.figure(figsize=(18, 12), facecolor="white")
fig.suptitle(
    "NY State Medicaid Expenditure — Four-Model Forecast (FY 2026–2035)",
    fontsize=15, fontweight="bold", y=0.98, color="#1A1A1A"
)
axes = fig.subplots(2, 2)

# shared helper
def shade_history(ax):
    ax.axvspan(years[0] - 0.5, 2025.5, color=C["history"], alpha=0.25, zorder=0)
    ax.axvline(2025.5, color=C["divider"], linestyle="--", linewidth=1.2, zorder=1)
    ax.text(2025.6, ax.get_ylim()[1] * 0.97, "Forecast →",
            color=C["divider"], fontsize=7.5, va="top")

# ── Panel 1 (top-left): Conservative models with 95% CI ─────────────────────
ax = axes[0, 0]

# History band
ax.fill_between(years, 0, y, alpha=0.08, color=C["actual"], zorder=0)

# Holt-Winters CI
ax.fill_between(FORECAST_YEARS, hw_ci_lower, hw_ci_upper,
                alpha=0.15, color=C["Holt-Winters"], zorder=2)
ax.plot(FORECAST_YEARS, fc_future_hw, "-", color=C["Holt-Winters"],
        linewidth=2.2, label=f"Holt-Winters  MAPE={mape_hw:.1f}%", zorder=4)

# Prophet CI
ax.fill_between(FORECAST_YEARS, prophet_ci_lower, prophet_ci_upper,
                alpha=0.15, color=C["Prophet"], zorder=2)
ax.plot(FORECAST_YEARS, fc_future_prophet, "-", color=C["Prophet"],
        linewidth=2.2, label=f"Prophet         MAPE={mape_prophet:.1f}%", zorder=4)

# ARIMA (flat — show as dashed)
ax.plot(FORECAST_YEARS, fc_future_arima, "--", color=C["ARIMA"],
        linewidth=2, label=f"ARIMA             MAPE={mape_arima:.1f}%", zorder=4)

# Actuals
ax.scatter(years, y, color=C["actual"], s=52, zorder=6, label="Actual (NY)", linewidths=0.5)

ax.yaxis.set_major_formatter(DOLLAR_FMT)
ax.set_xlim(years[0] - 0.5, 2035.5)
ax.set_ylim(45, 170)
shade_history(ax)
ax.set_title("Conservative Models — ARIMA, Holt-Winters, Prophet")
ax.set_xlabel("Fiscal Year")
ax.set_ylabel("Expenditures ($Billions)")
ax.legend(loc="upper left", handlelength=1.8)

# ── Panel 2 (top-right): Model performance — clean grouped horizontal bars ───
ax = axes[0, 1]

model_names = [m[0] for m in models_meta]
maes   = np.array([m[2] for m in models_meta])
rmses  = np.array([m[3] for m in models_meta])
mapes  = np.array([m[4] for m in models_meta])
colors = [C[n] for n in model_names]

# Normalize MAPE to dollars scale for side-by-side display (just show as separate group)
y_pos = np.arange(len(model_names))
h = 0.26

bars_mae  = ax.barh(y_pos + h,  maes,  h, color=colors, alpha=0.90, label="MAE ($B)")
bars_rmse = ax.barh(y_pos,      rmses, h, color=colors, alpha=0.55, label="RMSE ($B)", hatch="///")
bars_mape = ax.barh(y_pos - h,  mapes, h, color=colors, alpha=0.35, label="MAPE (%)")

# value labels
for bar, val, unit in [(bars_mae, maes, "B"), (bars_rmse, rmses, "B"), (bars_mape, mapes, "%")]:
    for b, v in zip(bar, val):
        ax.text(b.get_width() + 0.25, b.get_y() + b.get_height() / 2,
                f"{v:.1f}{unit}", va="center", ha="left", fontsize=7.5, color="#333333")

ax.set_yticks(y_pos)
ax.set_yticklabels(model_names, fontsize=9.5)
ax.invert_yaxis()
ax.set_xlabel("Error Magnitude  ($ Billions  |  %)")
ax.set_title("Model Accuracy — Test Set FY 2023–2025\n(MAE / RMSE in $B · MAPE in %)")
ax.set_xlim(0, max(rmses) * 1.30)
ax.legend(loc="lower right", fontsize=8)
ax.grid(axis="x", color="#E0E0E0", linewidth=0.7)
ax.grid(axis="y", visible=False)

# ── Panel 3 (bottom-left): Holt-Winters — recommended model deep-dive ────────
ax = axes[1, 0]

ax.fill_between(years, 0, y, alpha=0.08, color=C["actual"], zorder=0)

# 95% CI band
ax.fill_between(FORECAST_YEARS, hw_ci_lower, hw_ci_upper,
                alpha=0.22, color=C["Holt-Winters"], zorder=2, label="95% CI")

# Forecast line
ax.plot(FORECAST_YEARS, fc_future_hw, "-o", color=C["Holt-Winters"],
        linewidth=2.4, markersize=5, zorder=4, label="Holt-Winters forecast")

# Historical fitted values
hw_fitted = hw_full.fittedvalues
ax.plot(years, hw_fitted, "--", color=C["Holt-Winters"],
        linewidth=1.4, alpha=0.6, zorder=3, label="Historical fit")

# Actual dots
ax.scatter(years, y, color=C["actual"], s=55, zorder=6,
           label="Actual (NY)", linewidths=0.5)

# Annotate every other forecast year
for i, (yr, val) in enumerate(zip(FORECAST_YEARS, fc_future_hw)):
    if i % 2 == 0:
        ax.text(yr, val + 1.8, f"${val:,.1f}B",
                ha="center", va="bottom", fontsize=7.2, color=C["Holt-Winters"])

ax.yaxis.set_major_formatter(DOLLAR_FMT)
ax.set_xlim(years[0] - 0.5, 2035.5)
ax.set_ylim(45, 170)
shade_history(ax)
ax.set_title("Holt-Winters (Recommended) — Detailed Forecast with 95% CI")
ax.set_xlabel("Fiscal Year")
ax.set_ylabel("Expenditures ($Billions)")
ax.legend(loc="upper left")
ax.set_facecolor("#FAFAFA")

# ── Panel 4 (bottom-right): Grouped bar — key forecast milestones ────────────
ax = axes[1, 1]

milestones = [2026, 2030, 2035]
milestone_idx = [FORECAST_YEARS.index(y) for y in milestones]
n_models = len(models_meta)
n_ms = len(milestones)
bar_w = 0.18
offsets = np.linspace(-(n_models - 1) * bar_w / 2, (n_models - 1) * bar_w / 2, n_models)
x_ms = np.arange(n_ms)

for i, (name, cfg, mae, rmse, mape, fc) in enumerate(models_meta):
    vals = [fc[j] for j in milestone_idx]
    bars = ax.bar(x_ms + offsets[i], vals, bar_w,
                  color=C[name], alpha=0.88, label=name, zorder=3)
    for b, v in zip(bars, vals):
        ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 1.5,
                f"${v:,.0f}B", ha="center", va="bottom", fontsize=6.8,
                color="#333333", rotation=35)

ax.set_xticks(x_ms)
ax.set_xticklabels([str(m) for m in milestones], fontsize=11)
ax.yaxis.set_major_formatter(DOLLAR_FMT)
ax.set_title("Forecast at Key Milestones — All Four Models")
ax.set_xlabel("Fiscal Year")
ax.set_ylabel("Expenditures ($Billions)")
ax.legend(loc="upper left", ncol=2, fontsize=8)
ax.set_facecolor("#FAFAFA")

# ── Final layout ─────────────────────────────────────────────────────────────
plt.tight_layout(rect=[0, 0, 1, 0.97], h_pad=3.5, w_pad=3.0)
plt.savefig("all_models_results.png", dpi=160, bbox_inches="tight",
            facecolor="white", edgecolor="none")
plt.show()
print("\nChart saved → all_models_results.png")

# ---------------------------------------------------------------------------
# 9.  Export forecast table to CSV
# ---------------------------------------------------------------------------

forecast_df = pd.DataFrame({"Year": FORECAST_YEARS})
for name, cfg, mae, rmse, mape, fc in models_meta:
    forecast_df[name] = [f"${v:.2f}B" for v in fc]

forecast_df.to_csv("forecast_table.csv", index=False)
print("Forecast table saved → forecast_table.csv")

# Print final clean table
print("\n" + "=" * 95)
print("FINAL SUMMARY TABLE")
print("=" * 95)
print(f"\n{'Metric':<12} {'Poly Reg':>14} {'ARIMA':>14} {'Holt-Winters':>14} {'Prophet':>14}")
print("-" * 70)
print(f"{'MAE ($B)':<12} {mae_reg:>13.3f}  {mae_arima:>13.3f}  {mae_hw:>13.3f}  {mae_prophet:>13.3f}")
print(f"{'RMSE ($B)':<12} {rmse_reg:>13.3f}  {rmse_arima:>13.3f}  {rmse_hw:>13.3f}  {rmse_prophet:>13.3f}")
print(f"{'MAPE (%)':<12} {mape_reg:>13.2f}  {mape_arima:>13.2f}  {mape_hw:>13.2f}  {mape_prophet:>13.2f}")
print()
print(forecast_df.to_string(index=False))
