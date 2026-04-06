"""
Medicaid Expenditure Regression Model — New York State
Loads MFCU statistical data (FY 2013-2025) and predicts NY Total Medicaid Expenditures
using Linear, Polynomial, and Multi-feature regression, forecasting through FY 2036.
"""

import glob
import re
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures, StandardScaler
from sklearn.pipeline import make_pipeline
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# 1. Load all Excel files
# ---------------------------------------------------------------------------

NUMERIC_COLS = [
    "Total Investigations",
    "Fraud Investigations",
    "Abuse/Neglect Investigations",
    "Total Indictments",
    "Fraud Indictments",
    "Abuse/Neglect Indictments",
    "Total Convictions",
    "Fraud Convictions",
    "Abuse/Neglect Convictions",
    "Civil Settlements and Judgments",
    "Total Recoveries",
    "Total Criminal Recoveries",
    "Civil Recoveries Global",
    "Civil Recoveries Other",
    "MFCU Grant Expenditures",
    "Total Medicaid Expenditures",
    "Staff On Board",
]

NY_ALIASES = {"new york", "new york state", "ny"}


def load_year(filepath):
    """Return a DataFrame of state-level rows for one fiscal year file."""
    year_match = re.search(r"FY_(\d{4})", filepath)
    year = int(year_match.group(1))

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    ncols = ws.max_column  # 17 for 2013-2014, 18 for 2015+

    def to_num(v):
        return v if isinstance(v, (int, float)) else None

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        state = row[0]
        if state is None or not isinstance(state, str):
            continue
        if len(state) > 30:
            continue
        if state.strip().upper().replace(":", "").strip() in ("TOTAL", "GRAND TOTAL"):
            state = "Total"

        if ncols == 17:
            nums = [to_num(row[i]) for i in range(1, 14)]
            nums.append(None)
            nums += [to_num(row[i]) for i in range(14, 17)]
        else:
            nums = [to_num(row[i]) for i in range(1, 18)]

        rows.append([year, state] + nums)

    df = pd.DataFrame(rows, columns=["FY", "State"] + NUMERIC_COLS)
    return df


files = sorted(glob.glob("FY_*_MFCU_Statistical_Chart*.xlsx"))
all_data = pd.concat([load_year(f) for f in files], ignore_index=True)

# Filter to New York state rows only
ny_mask = all_data["State"].str.strip().str.lower().isin(NY_ALIASES)
ny_data = all_data[ny_mask].copy()

if ny_data.empty:
    raise ValueError("No New York rows found. Check state name spelling in the Excel files.")

ny_data = ny_data.sort_values("FY").reset_index(drop=True)

print(f"Loaded {len(files)} years | {len(ny_data)} NY state-year rows")
print(ny_data[["FY", "Total Medicaid Expenditures", "MFCU Grant Expenditures", "Staff On Board"]].to_string(index=False))

# ---------------------------------------------------------------------------
# 2. NY time-series regression (target = Total Medicaid Expenditures)
# ---------------------------------------------------------------------------

ts = ny_data.dropna(subset=["Total Medicaid Expenditures"]).reset_index(drop=True)
X_year = ts[["FY"]].values
y = ts["Total Medicaid Expenditures"].values / 1e9  # express in billions

FUTURE_YEARS = list(range(2026, 2037))  # 2026 through 2036
X_future = np.array(FUTURE_YEARS).reshape(-1, 1)

# ---- 2a. Simple Linear Regression ----
lr = LinearRegression()
lr.fit(X_year, y)
y_pred_lr = lr.predict(X_year)
y_future_lr = lr.predict(X_future)
r2_lr = r2_score(y, y_pred_lr)
mae_lr = mean_absolute_error(y, y_pred_lr)

# ---- 2b. Polynomial Regression (degree 2) ----
poly2 = make_pipeline(PolynomialFeatures(degree=2), LinearRegression())
poly2.fit(X_year, y)
y_pred_p2 = poly2.predict(X_year)
y_future_p2 = poly2.predict(X_future)
r2_p2 = r2_score(y, y_pred_p2)
mae_p2 = mean_absolute_error(y, y_pred_p2)

# ---- 2c. Polynomial Regression (degree 3) ----
poly3 = make_pipeline(PolynomialFeatures(degree=3), LinearRegression())
poly3.fit(X_year, y)
y_pred_p3 = poly3.predict(X_year)
y_future_p3 = poly3.predict(X_future)
r2_p3 = r2_score(y, y_pred_p3)
mae_p3 = mean_absolute_error(y, y_pred_p3)

print("\n=== NY Time-Series Regression Results ===")
print(f"{'Model':<20} {'R2':>8} {'MAE ($B)':>12}")
print("-" * 42)
print(f"{'Linear':<20} {r2_lr:>8.4f} {mae_lr:>12.2f}")
print(f"{'Polynomial (deg 2)':<20} {r2_p2:>8.4f} {mae_p2:>12.2f}")
print(f"{'Polynomial (deg 3)':<20} {r2_p3:>8.4f} {mae_p3:>12.2f}")

print("\n=== Predictions (NY Total Medicaid Expenditures, $Billions) ===")
print(f"{'Year':<8} {'Linear':>12} {'Poly-2':>12} {'Poly-3':>12}")
print("-" * 48)
for yr, v1, v2, v3 in zip(FUTURE_YEARS, y_future_lr, y_future_p2, y_future_p3):
    print(f"{yr:<8} {v1:>12.1f} {v2:>12.1f} {v3:>12.1f}")

# ---------------------------------------------------------------------------
# 3. Multi-feature regression (NY panel)
# ---------------------------------------------------------------------------

FEATURE_COLS = [
    "Total Investigations",
    "Total Indictments",
    "Total Convictions",
    "Total Recoveries",
    "MFCU Grant Expenditures",
    "Staff On Board",
]
TARGET = "Total Medicaid Expenditures"

panel = ny_data.dropna(subset=FEATURE_COLS + [TARGET]).copy()
panel[FEATURE_COLS + [TARGET]] = panel[FEATURE_COLS + [TARGET]].apply(pd.to_numeric, errors="coerce")
panel = panel.dropna(subset=FEATURE_COLS + [TARGET])

X_panel = panel[FEATURE_COLS].values
y_panel = panel[TARGET].values / 1e9

# Train on FY 2013-2022, test on FY 2023-2025
train_mask = panel["FY"] <= 2022
X_train, X_test = X_panel[train_mask], X_panel[~train_mask]
y_train, y_test = y_panel[train_mask], y_panel[~train_mask]

print(f"\n=== Multi-Feature Linear Regression (NY State) ===")

if X_train.shape[0] < 2 or X_test.shape[0] < 1:
    print("Insufficient NY data for multi-feature train/test split — skipping.")
    r2_mlr = mae_mlr = rmse_mlr = None
    coef_df = pd.DataFrame({"Feature": FEATURE_COLS, "Coefficient": [0] * len(FEATURE_COLS)})
else:
    scaler = StandardScaler()
    X_train_s = scaler.fit_transform(X_train)
    X_test_s = scaler.transform(X_test)

    mlr = LinearRegression()
    mlr.fit(X_train_s, y_train)
    y_pred_test = mlr.predict(X_test_s)

    r2_mlr = r2_score(y_test, y_pred_test)
    mae_mlr = mean_absolute_error(y_test, y_pred_test)
    rmse_mlr = np.sqrt(mean_squared_error(y_test, y_pred_test))

    print(f"Train: FY 2013-2022 ({train_mask.sum()} rows)  |  Test: FY 2023-2025 ({(~train_mask).sum()} rows)")
    print(f"R2: {r2_mlr:.4f} | MAE: ${mae_mlr:.2f}B | RMSE: ${rmse_mlr:.2f}B")

    coef_df = pd.DataFrame({"Feature": FEATURE_COLS, "Coefficient": mlr.coef_})
    print("\nFeature Coefficients (scaled):")
    print(coef_df.sort_values("Coefficient", ascending=False).to_string(index=False))

# ---------------------------------------------------------------------------
# 4. Plot results
# ---------------------------------------------------------------------------

fig, axes = plt.subplots(2, 2, figsize=(16, 11))
fig.suptitle("New York State Medicaid Expenditure Regression Analysis (FY 2013–2025, Forecast to 2036)",
             fontsize=13, fontweight="bold")

years_all = np.arange(ts["FY"].min(), 2037)
X_all = years_all.reshape(-1, 1)

# -- Plot 1: Time-series regression lines --
ax = axes[0, 0]
ax.scatter(ts["FY"], y, color="black", zorder=5, label="Actual (NY)", s=60)
ax.plot(years_all, lr.predict(X_all), "--", color="steelblue", label=f"Linear (R2={r2_lr:.3f})")
ax.plot(years_all, poly2.predict(X_all), "-", color="darkorange", label=f"Poly-2 (R2={r2_p2:.3f})")
ax.plot(years_all, poly3.predict(X_all), "-.", color="green", label=f"Poly-3 (R2={r2_p3:.3f})")
ax.axvline(2025.5, color="gray", linestyle=":", linewidth=1, label="Forecast →")
ax.set_title("NY Total Medicaid Expenditures — Regression & Forecast")
ax.set_xlabel("Fiscal Year")
ax.set_ylabel("Expenditures ($Billions)")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}B"))
ax.legend(fontsize=8)
ax.grid(True, alpha=0.3)

# -- Plot 2: Forecast bar chart (best model by R2) --
ax = axes[0, 1]
best_r2 = max(r2_lr, r2_p2, r2_p3)
if best_r2 == r2_p3:
    best_future, best_label = y_future_p3, "Poly-3"
elif best_r2 == r2_p2:
    best_future, best_label = y_future_p2, "Poly-2"
else:
    best_future, best_label = y_future_lr, "Linear"

hist_years = list(ts["FY"])
hist_vals = list(y)
bars_hist = ax.bar(hist_years, hist_vals, color="steelblue", alpha=0.7, label="Actual (NY)")
bars_pred = ax.bar(FUTURE_YEARS, best_future, color="darkorange", alpha=0.8, label=f"Forecast ({best_label})")
ax.set_title(f"NY Expenditure Forecast to 2036 (Best Model: {best_label})")
ax.set_xlabel("Fiscal Year")
ax.set_ylabel("Expenditures ($Billions)")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}B"))
ax.legend()
ax.grid(True, alpha=0.3, axis="y")
# Label every other forecast bar to avoid crowding
for i, (bar, val) in enumerate(zip(bars_pred, best_future)):
    if i % 2 == 0:
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                f"${val:,.1f}B", ha="center", va="bottom", fontsize=7, rotation=45)

# -- Plot 3: Historical NY trend with all three model fits --
ax = axes[1, 0]
ax.scatter(ts["FY"], y, color="black", zorder=5, s=50, label="Actual")
ax.plot(ts["FY"], y_pred_lr, "o--", color="steelblue", markersize=4, label=f"Linear fit")
ax.plot(ts["FY"], y_pred_p2, "s-", color="darkorange", markersize=4, label=f"Poly-2 fit")
ax.plot(ts["FY"], y_pred_p3, "^-.", color="green", markersize=4, label=f"Poly-3 fit")
ax.set_title("NY Historical Fit Comparison (FY 2013–2025)")
ax.set_xlabel("Fiscal Year")
ax.set_ylabel("Expenditures ($Billions)")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}B"))
ax.legend(fontsize=8)
ax.grid(True, alpha=0.3)

# -- Plot 4: Feature importance (coefficients) or fallback text --
ax = axes[1, 1]
if r2_mlr is not None:
    sorted_coef = coef_df.sort_values("Coefficient")
    colors = ["tomato" if c < 0 else "mediumseagreen" for c in sorted_coef["Coefficient"]]
    ax.barh(sorted_coef["Feature"], sorted_coef["Coefficient"], color=colors)
    ax.axvline(0, color="black", linewidth=0.8)
    ax.set_title("Feature Coefficients (Multi-Feature Model, Scaled)\nNY State")
    ax.set_xlabel("Coefficient Value")
    ax.grid(True, alpha=0.3, axis="x")
else:
    ax.text(0.5, 0.5, "Insufficient data\nfor multi-feature model",
            ha="center", va="center", transform=ax.transAxes, fontsize=12)
    ax.set_title("Feature Coefficients — NY State")

plt.tight_layout()
plt.savefig("regression_results.png", dpi=150, bbox_inches="tight")
plt.show()
print("\nPlot saved to regression_results.png")

# ---------------------------------------------------------------------------
# 5. Summary table
# ---------------------------------------------------------------------------

print(f"\n=== NY State Forecast Summary (Best Model: {best_label}) ===")
print(f"{'Year':<6} {'Predicted ($B)':>16} {'Change vs Prior':>16}")
prev = y[-1]  # last known actual (FY 2025)
for yr, val in zip(FUTURE_YEARS, best_future):
    change = val - prev
    pct = (change / prev) * 100
    print(f"{yr:<6} ${val:>13,.1f}   {'+' if change>=0 else ''}{change:>8.1f}B ({pct:+.1f}%)")
    prev = val
